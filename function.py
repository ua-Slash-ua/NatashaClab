import random
import traceback
import json
import logging
import colorlog
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# URL для входу на сайт
logining_url = 'https://www.natashaclub.com/member.php'
base_url = 'https://www.natashaclub.com/'
send_message_url = 'https://www.natashaclub.com/compose.php?ID='
start_search = 'https://www.natashaclub.com/search_result.php?p_per_page=10&online_only=on&Sex=female&LookingFor=male&DateOfBirth_start=35&DateOfBirth_end=75&Region[]=3&Region[]=4&Region[]=7&CityST=0&City=&TrustLevel=0&Registered=0&LastOnline=0&LanguageSkills1=0&&page='
com_url = 'https://www.natashaclub.com/cc.php'
outbox_url_base = 'https://www.natashaclub.com/outbox.php'
ajax_url = 'https://www.natashaclub.com/ajax.action.php'
# ANSI escape sequences для кольорів у словнику
COLORS = {
    'RESET' : "\033[0m",
    'BOLD' : "\033[1m",
    'RED' : "\033[31m",
    'GREEN' : "\033[32m",
    'YELLOW' : "\033[33m",
    'BLUE' : "\033[34m",
    'MAGENTA' : "\033[35m",
    'CYAN' : "\033[36m",
    'WHITE' : "\033[37m",
    'MY':"\033[1m"
}
payload_authorization = {}
payload_text = {}
start_words = ['Dear','Honey','Darling']

# Налаштування логування
logging.basicConfig(level=logging.INFO,
                    format='%(levelname)s - %(message)s')
logging.basicConfig(level=logging.WARNING,
                    format='%(levelname)s - %(message)s')
logging.basicConfig(level=logging.ERROR,
                    format='%(levelname)s - %(message)s')
logging.basicConfig(level=logging.WARN,
                    format='%(levelname)s - %(message)s')

def load_settings(path_to_account):
    with open(path_to_account, 'r', encoding='utf-8') as fj:
        settings = json.load(fj)
    return settings

def load_name(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            # Зчитуємо дані з JSON файлу
            names = json.load(f)  # Завантажуємо вміст файлу у змінну names
            return names
    except:
        print("Файл порожній або має помилковий формат.")
        return {}

def write_html(soup):
    with open('br.html','w', encoding='utf-8') as file:
        file.write(str(soup))
    logging.info('Сторінку записано!')

def write_name(path, data):
    with open(path,'w', encoding='utf-8') as f:
        json.dump(data,f,indent=4)
    logging.info('Імена оновлено!!!')

def count_pages(session):
    response_page_search = session.post(f'{start_search}1')
    soup_page_search = BeautifulSoup(response_page_search.text, 'lxml')
    n = soup_page_search.find('div', class_='DataDiv').find_all(string=True)
    count_pages = [num for num in n if 'found. Pages: ' in num]
    return int(count_pages[0][:3])//10

def authorization_on_account(name,session_in):
        # authorization
        response_authorization = session_in.post(logining_url, payload_authorization[name])
        response_authorization_audit = session_in.get(logining_url)
        soup_authorization_audit = BeautifulSoup(response_authorization_audit.text, 'lxml')
        return True if soup_authorization_audit.title.text == 'Control Panel' else False

def write_me(session,username_man):
    pay = {'filterID': username_man}
    response_outbox = session.post(outbox_url_base,pay)
    soup_outbox = BeautifulSoup(response_outbox.text, 'lxml')
    result = soup_outbox.find('td', class_ = 'panel').text.split('\n')
    for line in result:
        if line.startswith('\xa0Messages in Outbox:'):
            num_list = int(line[line.rindex('Outbox:')+7:line.index('total')].strip())
            if num_list ==0:
                return True
            else:
                return False

def get_link_on_exel(folder_path: str) -> dict:
    # Завантажуємо книгу з Excel-файлу
    wb = load_workbook(folder_path, data_only=True)
    ws = wb.active

    # Ініціалізуємо словник для зберігання гіперпосилань
    man_info = {}

    # Проходимося по всіх комірках активного листа
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink:
                # Зберігаємо комірку та гіперпосилання
                if '.' in str(cell.value):
                    cell.value = str(cell.value)[:str(cell.value).index('.')]
                    man_info[cell.value] = cell.hyperlink.target
                else:
                    man_info[cell.value] = cell.hyperlink.target
    return man_info

def get_link_with_txt(path_folder: str) -> list[str]:
    data = []
    with open(path_folder, 'r', encoding='utf-8') as file:
        data_list = file.read().split('\n')
        for line in data_list:
            if line.startswith('https://www.natashaclub.com/profile.php?ID='):
                data.append(line)
    return data

def get_payload_account(path_to_account):
    with open(path_to_account, 'r', encoding='utf-8') as fj:
        accounts = json.load(fj)
        for account in accounts:
            payload_authorization[account] =accounts[account]['authorization']
            payload_text[account] = accounts[account]['text']
    return payload_authorization,payload_text

def colored_log(level, msg:str = '', e:Exception = None):
    color = COLORS['RESET']  # за замовчуванням без кольору
    if level == logging.DEBUG:
        color = COLORS['BOLD']
        print(f"{color}{msg}{COLORS['RESET']}")
    elif level == logging.INFO:
        color = COLORS['GREEN']
        print(f">>> {color}{msg}{COLORS['RESET']}")
    elif level == logging.WARN:
        color = COLORS['CYAN']
        print(f"{color}{msg}{COLORS['RESET']}")
    elif level == logging.WARNING:
        color = COLORS['YELLOW']
        print(f"{color}{msg}{COLORS['RESET']}")
    elif level == logging.ERROR:
        # Отримуємо тип виключення
        exception_type = type(e).__name__

        # Отримуємо номер рядка та інші деталі
        last_trace = traceback.extract_tb(e.__traceback__)[-1]  # Останній запис в стеку
        filename, lineno, funcname, text = last_trace  # Отримуємо деталі

        # Формуємо повідомлення про помилку
        error_message = f"В {lineno} рядку відбулася помилка _{exception_type}_: {str(e)}"
        color = COLORS['RED']
        print(f"{color}{error_message}{COLORS['RESET']}")

def get_text(start_message,text_message):
    new_text = ''
    s = [ i for i in start_words if text_message.startswith(i)]
    if start_message in start_words:

        text_message = text_message[len(s[0])+1:] if s else text_message
        new_text = text_message.replace('NAME', start_message)
    else:
        new_text = text_message.replace('NAME', start_message)
    return new_text

def get_start_message(header_man: str, names, skip_send):
    name_man = header_man[:header_man.index(':')].strip().lower()
    new_name = ''
    # Залишає літери та пробіли
    for symbol in name_man:
        if symbol.isalpha() or symbol.isspace():
            new_name += symbol

    name_man_list = [name_man_part.strip() for name_man_part in new_name.split()]
    # Якщо ім'я не знайдено :
    # Вводимо ім'я і записуємо новий варіант до БД
    #
    for name_man_part in name_man_list:
        for name in names:
            # якщо ім'я знайдено і воно відоме тобто не дорівнює \ skip /
            if name_man_part == name and names[name] != 'skip':
                r_name = names[name].capitalize()
                return r_name
            # якщо ім'я знайдено і воно невідоме тобто, дорівнює \ skip /
            elif name_man_part == name and names[name] == 'skip':
                r_start_message = random.choice(start_words)
                return r_start_message
        if skip_send:
            return 'skip'
        # якщо ім'я знайдено
        else:
            # якщо ім'я одне
            if len(name_man_list) == 1:
                while True:
                    new_start_message = ''
                    start_message = input('Write man name:\n>>>').strip()
                    if start_message == 'skip':
                        names[name_man_list[0]] = 'skip'
                        r_start_message = random.choice(start_words)
                        return r_start_message
                    else:
                        for symbol in start_message:
                            if symbol.isalpha():
                                new_start_message += symbol
                        if len(new_start_message) < 2:
                            continue
                        names[name_man_list[0]] = new_start_message.lower()
                        break
                r_start_message = new_start_message.capitalize()
                return r_start_message
            # якщо імен більше 1
            else:
                r_start_message_list = []
                num_name = len(name_man_list)
                count_name = 1
                while count_name<=num_name:
                    new_start_message = ''
                    start_message = input(f'Write __ {count_name} __ man name:\n>>>').strip()
                    if start_message == 'skip':
                        names[name_man_list[count_name-1]] = 'skip'
                        r_start_message = random.choice(start_words)
                        r_start_message_list.append(r_start_message)
                    else:
                        for symbol in start_message:
                            if symbol.isalpha():
                                new_start_message += symbol
                        if len(new_start_message) < 2:
                            continue
                        else:
                            names[name_man_list[count_name-1]] = new_start_message.lower()
                            r_start_message = new_start_message.capitalize()
                            r_start_message_list.append(r_start_message)
                    count_name+=1
                # Перевірка чи є хоч одне нормальне ім'я, якщо є то відправляє якщо ні то відправляє \ start_words /

                r_start_message_list = [name_word for name_word in r_start_message_list if name_word not in start_words]
                if r_start_message_list:
                    return r_start_message_list[0]
                else:
                    r_start_message = random.choice(start_words)
                    return r_start_message


from openpyxl import load_workbook


def load_man_link(path_to_xlsx):
    data = []
    try:

        workbook = load_workbook(filename=path_to_xlsx, data_only=True)
        sheet = workbook.active

        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    data.append(cell.hyperlink.target)

    except FileNotFoundError:
        print(f"Файл не знайдено: {path_to_xlsx}")
    except Exception as e:
        print(f"Сталася помилка: {e}")
    finally:
        return data



